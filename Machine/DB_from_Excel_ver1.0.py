import sys
import os
import json
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QHBoxLayout,
    QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem
)
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def process_files(files):
    """
    指定されたExcelファイルを読み込み、D列とL列からサイクル数と放電容量を取得し、
    AU列〜BA列の4行目〜18行目の電極情報をまとめる。
    結果をDataFrameとして返す。
    """
    results = []
    for file in files:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        data = {}
        for col_letter, suffix in [('D', 'A'), ('L', 'B')]:
            col_idx = column_index_from_string(col_letter)
            cycle = ws.cell(row=3, column=col_idx).value
            capacity = None
            for r in range(3, ws.max_row + 2):
                if ws.cell(row=r, column=col_idx).value is None:
                    capacity = ws.cell(row=r - 1, column=col_idx).value
                    break
            data[f'cycle{suffix}'] = cycle
            data[f'capacity{suffix}'] = capacity

        # 電極情報をAU〜BA列、4〜18行目から取得
        start_col = column_index_from_string('AU')
        end_col = column_index_from_string('BA')
        electrode = []
        for r in range(4, 19):
            row_vals = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
            electrode.append(row_vals)
        data['electrode_info'] = json.dumps(electrode, ensure_ascii=False)
        data['filename'] = os.path.basename(file)
        results.append(data)

    return pd.DataFrame(results)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Cycle Capacity Manager")
        self.df_master = pd.DataFrame(
            columns=[
                'filename', 'cycleA', 'capacityA', 'cycleB', 'capacityB', 'electrode_info'
            ]
        )
        self.loaded_files = set()  # 読み込んだファイルパスを管理

        # ボタンレイアウト
        btn_layout = QHBoxLayout()
        self.btn_load = QPushButton("Load Excel Files")
        self.btn_load.clicked.connect(self.load_files)
        self.btn_export = QPushButton("Output CSV")
        self.btn_export.clicked.connect(self.export_csv)
        btn_layout.addWidget(self.btn_load)
        btn_layout.addWidget(self.btn_export)

        # テーブルウィジェット
        self.table = QTableWidget(0, len(self.df_master.columns))
        self.table.setHorizontalHeaderLabels(self.df_master.columns)

        # 全体レイアウト
        layout = QVBoxLayout()
        layout.addLayout(btn_layout)
        layout.addWidget(self.table)
        self.setLayout(layout)

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files",
            os.getcwd(),
            "Excel Files (*.xlsx *.xls *.xlsm *.csv)"
        )
        if not files:
            return
        # 未読み込みのファイルのみ抽出
        new_files = [f for f in files if f not in self.loaded_files]
        if not new_files:
            QMessageBox.information(self, "Info", "All selected files have already been loaded.")
            return
        try:
            df_new = process_files(new_files)
            # DataFrame に追加
            self.df_master = pd.concat([self.df_master, df_new], ignore_index=True)
            # 読み込んだファイルを登録
            self.loaded_files.update(new_files)
            self.update_table()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def update_table(self):
        self.table.setRowCount(len(self.df_master))
        for row_idx, (_, row) in enumerate(self.df_master.iterrows()):
            for col_idx, col in enumerate(self.df_master.columns):
                item = QTableWidgetItem(str(row[col]))
                self.table.setItem(row_idx, col_idx, item)

    def export_csv(self):
        if self.df_master.empty:
            QMessageBox.warning(self, "Warning", "No data to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save CSV",
            os.path.join(os.getcwd(), 'summary.csv'),
            "CSV Files (*.csv)"
        )
        if not path:
            return
        try:
            self.df_master.to_csv(path, index=False)
            QMessageBox.information(self, "Saved", f"CSV exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(800, 600)
    window.show()
    sys.exit(app.exec_())

# 実行方法:
# pip install pandas openpyxl pyqt5
# python excel_cycle_capacity_gui.py
