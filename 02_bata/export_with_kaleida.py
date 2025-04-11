
import openpyxl
from openpyxl import Workbook

def export_to_excel_with_kaleida(output_path, voltage, capacity_data_by_cycle):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "GraphData"

    # GraphDataに出力
    ws1.cell(row=1, column=1, value="Voltage")
    for i, cycle_num in enumerate(sorted(capacity_data_by_cycle.keys()), start=2):
        ws1.cell(row=1, column=i, value=f"Cycle {cycle_num}")

    for row_idx, v in enumerate(voltage, start=2):
        ws1.cell(row=row_idx, column=1, value=v)
        for col_idx, cycle_num in enumerate(sorted(capacity_data_by_cycle.keys()), start=2):
            cap_list = capacity_data_by_cycle[cycle_num]
            if row_idx - 2 < len(cap_list):
                ws1.cell(row=row_idx, column=col_idx, value=cap_list[row_idx - 2])

    # Kaleidaシート追加
    ws2 = wb.create_sheet(title="Kaleida")
    ws2.cell(row=1, column=1, value="Voltage")
    for idx, cycle_num in enumerate(sorted(capacity_data_by_cycle.keys()), start=2):
        ws2.cell(row=1, column=idx, value=f"Cycle {cycle_num}")

    for row_idx, v in enumerate(voltage, start=2):
        ws2.cell(row=row_idx, column=1, value=v)
        for col_idx, cycle_num in enumerate(sorted(capacity_data_by_cycle.keys()), start=2):
            cap_list = capacity_data_by_cycle[cycle_num]
            if row_idx - 2 < len(cap_list):
                ws2.cell(row=row_idx, column=col_idx, value=cap_list[row_idx - 2])

    wb.save(output_path)
